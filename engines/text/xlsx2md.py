#!/usr/bin/env python3
"""
xlsx2md — Excel (.xlsx) to Markdown. Sibling to pdf2md.py/docx2md.py; shares
the same CLI contract via common.py.

Like docx, .xlsx is already structured and text-native -- no OCR, no
possibility of the OCR/LLM spelling-decision defect this project exists to
prevent. The engineering concerns are specific to spreadsheets:

  - read CACHED VALUES (data_only=True), never raw formula text -- a cell
    showing "=SUM(B2:B10)" as its "value" would misrepresent the actual
    reported figure, not just be inconvenient to read.
  - convert EVERY sheet, not just the first -- financial workbooks commonly
    split Balance Sheet / P&L / Notes across sheets.
  - a formula cell with NO cached value (the workbook was generated
    programmatically and never opened/saved in Excel, so no cached result
    exists) is flagged explicitly as "[unresolved formula]", never
    silently rendered blank -- "we can't determine this" and "this cell is
    genuinely blank" are different facts and must stay distinguishable.
  - hidden sheets/rows/columns are included by default (openpyxl's normal
    iteration doesn't skip them) -- the conservative "never silently drop
    content" choice, consistent with the rest of this project.
  - merged cells: openpyxl's own structure genuinely only stores a value at
    the top-left cell of a merged range (every other cell in the range is
    None) -- unlike docx, this is the file's real internal structure, not a
    detection heuristic, so leaving non-top-left cells blank is accurate,
    not a guess.

Negative numbers render parenthesised, matching how pdf2md's PDF-derived
tables already show negatives, so downstream table-parsing code sees one
consistent sign convention regardless of source format. Output is the same
GFM pipe-table markdown syntax the other engines produce.
"""
import datetime
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from common import run


def format_cell(value, number_format: str = "General") -> str:
    """Reproduce the cell's own declared display convention rather than
    imposing one -- confirmed a real defect: blanket thousands-separator
    formatting turned a plain "2022" year-header cell into "2,022", which
    doesn't match how the source spreadsheet itself displays it. A cell's
    number_format string (e.g. "#,##0.00", "yyyy", "General") is openpyxl's
    record of the file's own formatting intent; only apply thousands
    separators when the source format itself uses them."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        neg = value < 0
        magnitude = abs(value)
        use_commas = "," in (number_format or "")
        if float(magnitude).is_integer():
            s = f"{int(magnitude):,}" if use_commas else str(int(magnitude))
        else:
            s = f"{magnitude:,.2f}" if use_commas else f"{magnitude:.2f}"
        return f"({s})" if neg else s
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip().replace("|", "\\|")


def merged_non_owner_cells(ws) -> set[tuple[int, int]]:
    """(row, col) positions inside a merged range that are NOT the top-left
    owning cell -- these hold no value in the file itself (see module
    docstring) and are rendered blank."""
    blanks = set()
    for merged_range in ws.merged_cells.ranges:
        owner = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != owner:
                    blanks.add((row, col))
    return blanks


def sheet_to_markdown(ws_values, ws_formulas) -> str:
    blanks = merged_non_owner_cells(ws_values)
    max_col = ws_values.max_column
    rows = []
    for row_idx in range(1, ws_values.max_row + 1):
        cells = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in blanks:
                cells.append("")
                continue
            v_cell = ws_values.cell(row=row_idx, column=col_idx)
            formula = ws_formulas.cell(row=row_idx, column=col_idx).value
            unresolved = v_cell.value is None and isinstance(formula, str) and formula.startswith("=")
            cells.append("[unresolved formula]" if unresolved else format_cell(v_cell.value, v_cell.number_format))
        rows.append(cells)

    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return ""

    lines = ["| " + " | ".join(rows[0]) + " |",
             "|" + "|".join("---" for _ in rows[0]) + "|"]
    for row_cells in rows[1:]:
        lines.append("| " + " | ".join(row_cells) + " |")
    return "\n".join(lines)


def to_markdown(input_path: str) -> str:
    import openpyxl

    wb_values = openpyxl.load_workbook(input_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(input_path, data_only=False)

    out = []
    for sheet_name in wb_values.sheetnames:
        table_md = sheet_to_markdown(wb_values[sheet_name], wb_formulas[sheet_name])
        if table_md:
            out.append(f"## {sheet_name}")
            out.append(table_md)

    text = "\n\n".join(out)
    if len(text.strip()) < 10:
        raise ValueError(f"produced almost no output ({len(text.strip())} chars) -- "
                          f"this .xlsx may be empty, corrupted, or in a form openpyxl can't read")
    return text


def _convert(args, log):
    log(f"[xlsx2md] reading {args.input}")
    out = to_markdown(args.input)
    log(f"[xlsx2md] extracted {len(out)} chars")
    return out


if __name__ == "__main__":
    run("xlsx2md", "Excel (.xlsx) to Markdown.", _convert)
